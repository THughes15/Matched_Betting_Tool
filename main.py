import os

import tasks
from pyperclip import copy
from tasks import *
from tkinter import Toplevel
import tkinter as tk
from tkinter import ttk
from tkinter import Frame
from tkinter import messagebox
from tkcalendar import DateEntry
from datetime import date
from openpyxl import load_workbook


class App(tk.Tk):

    # __init__ function for class tkinterApp
    def __init__(self, *args, **kwargs):
        # __init__ function for class Tk
        tk.Tk.__init__(self, *args, **kwargs)

        # Configure the root window
        self.title('Matched Betting Tool')
        self.geometry('427x180')

        # Creating a Container
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # initializing frames to an empty array
        self.frames = {}
        for F in (StartPage, Quali, Free):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        # Check if Google Tasks is Setup
        if os.path.exists('config.txt'):
            if os.path.exists('token.json'):
                tasks.initial()
            self.show_frame(StartPage)
        else:
            check = messagebox.askyesno('Google Tasks', 'Would you like to Integrate Google Tasks?')
            if check:
                f = open("config.txt", "a")
                f.write("Google Tasks: True")
                f.close()

                tasks.initial()
            else:
                f = open("config.txt", "a")
                f.write("Google Tasks: False")
                f.close()

    # Display the current frame passed as parameter
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()


# Home Page Frame
class StartPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        normal_inputs = Inputs(self, 'Normal')
        normal_inputs.grid(column=1, row=0, padx=20)

        # Menu Buttons
        menu = Menu(self, controller)
        menu.home_button['text'] = 'Quali Bet'
        menu.home_button['command'] = lambda: controller.show_frame(Quali)
        menu.grid(column=0, row=0)


# Qualification Bet Page Frame
class Quali(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        options = Inputs(self, 'Quali')
        options.grid(column=1, row=0, padx=20)

        menu = Menu(self, controller)
        menu.home_button['text'] = 'Home Page'
        menu.home_button['command'] = lambda: controller.show_frame(StartPage)
        menu.grid(column=0, row=0)


# Qualification Bet Page Frame
class Free(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        # Create Free Bets Frame
        bets_frame = Frame(self)
        label = tk.Text(bets_frame, width=35, height=7, font=("Helvetica", 12))
        label.grid(column=0, row=0)
        bets_frame.grid(column=1, row=0)

        # Create Submit Frame
        submit_frame = Frame(self)
        self.button = ttk.Button(submit_frame, text='Refresh', default='active', padding=5)
        self.button['command'] = lambda: insert_bets()
        self.button.grid(column=0, row=0)
        submit_frame.grid(column=1, row=1)

        menu = Menu(self, controller)
        menu.home_button['text'] = 'Home Page'
        menu.home_button['command'] = lambda: controller.show_frame(StartPage)
        menu.grid(column=0, row=0, rowspan=2)

        def insert_bets():
            label.delete(1.0, 'end')
            results = tasks.get_bets()
            for r in results:
                print(r)
                label.insert('end', f'{r["title"]}\n')


class Inputs(Frame):

    def __init__(self, parent, page):
        tk.Frame.__init__(self, parent)
        self.page = page

        # Create Bookie Frame
        bookie_frame = Frame(self)
        label = ttk.Label(bookie_frame, text='Bookmaker:')
        label.grid(column=0, row=0)

        self.bookie = tk.Entry(bookie_frame)
        self.bookie.grid(column=0, row=1)

        # Create Exchange Frame
        exchange_frame = Frame(self)
        label = ttk.Label(exchange_frame, text='Exchange:')
        label.grid(column=0, row=0)

        self.exchange = ttk.Combobox(exchange_frame, width=17)
        self.exchange['values'] = ('Smarkets',
                                   'Matchbook',
                                   'Betfair',
                                   'Betdaq')
        self.exchange.current(0)
        self.exchange.grid(column=0, row=1)

        # Place Details Frame
        details_frame = Frame(self)
        label = ttk.Label(details_frame, text='Details:')
        label.grid(column=0, row=0)

        self.details = tk.Entry(details_frame)
        self.details.grid(column=0, row=1)

        # Create Profit Frame
        profit_frame = Frame(self)
        label = ttk.Label(profit_frame, text='Profit/Loss:')
        label.grid(column=0, row=0)

        self.profit = tk.Entry(profit_frame)
        self.profit.grid(column=0, row=1)

        # Create Valid Frame
        valid_frame = Frame(self)
        label = ttk.Label(valid_frame, text='Available From:')
        label.grid(column=0, row=0)

        self.valid = DateEntry(valid_frame, selectmode='day', width=17, locale='en_UK', date_pattern='dd/mm/yy')
        self.valid.grid(column=0, row=1)

        # Create Submit Frame
        submit_frame = Frame(self)
        self.button = ttk.Button(submit_frame, text='Submit', default='active', padding=5)
        self.button['command'] = self.button_submit
        self.button.grid(column=0, row=2)

        # Place Frames
        bookie_frame.grid(column=0, row=0, pady=10, padx=10)
        exchange_frame.grid(column=1, row=0, pady=10, padx=10)
        details_frame.grid(column=0, row=1, pady=10, padx=10)
        profit_frame.grid(column=1, row=1, pady=10, padx=10)

        if self.page == 'Quali':
            valid_frame.grid(column=0, row=2, padx=10, pady=10)
            submit_frame.grid(column=1, row=2, pady=10, padx=10)
        else:
            submit_frame.grid(column=0, row=2, pady=10, padx=10, columnspan=2)

    def button_submit(self):

        # Read Content From File
        try:
            wb = load_workbook(filename="Betting Tool Log.xlsx")
            if self.page == 'Quali':
                ws = wb['Quali']
            else:
                ws = wb['Normal']

            row = ws.max_row + 1
            get_date = date.today()
            today = get_date.strftime('%d/%m/%y')

            # Format Date Cell
            cell = ws.cell(column=2, row=row)
            cell.number_format = 'dd/mm/yy;@'
            cell.value = today

            # Format Bookie Cell
            cell = ws.cell(column=3, row=row)
            cell.value = self.bookie.get()

            # Format Exchange Cell
            cell = ws.cell(column=4, row=row)
            cell.value = self.exchange.get()

            # Format Details Cell
            cell = ws.cell(column=5, row=row)
            cell.value = self.details.get()

            # Format Profit Cell
            cell = ws.cell(column=6, row=row)
            cell.number_format = '??#,##0.00'
            cell.value = float(self.profit.get())

            # Format Valid Cell
            if self.page == 'Quali':
                cell = ws.cell(column=7, row=row)
                cell.value = self.valid.get()

                # Add to Google Tasks
                info = f'{self.bookie.get()}: {self.details.get()}'
                self.valid.configure(date_pattern='Y-mm-dd')
                due_date = self.valid.get() + 'T00:00:00.000Z'
                self.valid.configure(date_pattern='dd/mm/yy')
                tasks.add_task(info, due_date)

            wb.save("Betting Tool Log.xlsx")
            wb.close()

        except PermissionError:
            messagebox.showerror('Error', 'Log File Permission Denied!')
            raise Exception('Log File Permission Denied!')

        self.bookie.delete(0, 'end')
        self.details.delete(0, 'end')
        self.profit.delete(0, 'end')

        messagebox.showinfo(title='Information', message='Success')


# Menu Frame
class Menu(Frame):
    def __init__(self, ws, controller):
        Frame.__init__(self, ws)
        self.ws = ws

        self.home_button = ttk.Button(self)
        self.home_button.grid(row=0, column=0, padx=10, pady=5)

        button1 = ttk.Button(self, text="Open Log",
                             command=lambda: os.startfile('Betting Tool Log.xlsx'))
        button1.grid(row=1, column=0, padx=10, pady=5)

        button2 = ttk.Button(self, text="Calculator",
                             command=lambda: self.open_window())
        button2.grid(row=2, column=0, padx=10, pady=5)

        button3 = ttk.Button(self, text="Free Bets",
                             command=lambda: controller.show_frame(Free))
        button3.grid(row=3, column=0, padx=10, pady=5)

        button4 = ttk.Button(self, text="Close",
                             command=controller.destroy)
        button4.grid(row=4, column=0, padx=10, pady=5)

    def open_window(self):
        window = Window(self)
        window.grab_set()


# Calculator Window
class Window(Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.geometry('295x345')
        self.title('Calculator')

        self.optimal = 0

        # Bet Type Selection
        type_frame = Frame(self)
        label = ttk.Label(type_frame, text='Bet Type:')
        label.grid(column=0, row=0)
        self.bet_type = ttk.Combobox(type_frame, width=17)
        self.bet_type['values'] = ('Qualification',
                                   'Free Bet (SNR)',
                                   'Free Bet (SR)')
        self.bet_type.current(0)
        self.bet_type.grid(column=0, row=1)

        # Back Stake Info
        frame1 = Frame(self)
        label = ttk.Label(frame1, text='Back Stake (??):')
        label.grid(column=0, row=0)
        self.stake = ttk.Entry(frame1)
        self.stake.grid(column=0, row=1)

        # Back Odds Info
        frame2 = Frame(self)
        label = ttk.Label(frame2, text='Back Odds:')
        label.grid(column=0, row=0)
        self.back_odds = ttk.Entry(frame2)
        self.back_odds.grid(column=0, row=1)

        # Lay Odds Info
        frame3 = Frame(self)
        label = ttk.Label(frame3, text='Lay Odds:')
        label.grid(column=0, row=0)
        self.lay_odds = ttk.Entry(frame3)
        self.lay_odds.grid(column=0, row=1)

        # Lay Commission Info
        frame4 = Frame(self)
        label = ttk.Label(frame4, text='Commission (%):')
        label.grid(column=0, row=0)
        self.commission = ttk.Entry(frame4)
        self.commission.insert('end', '0')
        self.commission.grid(column=0, row=1)

        type_frame.grid(column=0, row=0, columnspan=2, padx=10, pady=10)
        frame1.grid(column=0, row=1, padx=10, pady=10)
        frame2.grid(column=1, row=1, padx=10, pady=10)
        frame3.grid(column=0, row=2, padx=10, pady=10)
        frame4.grid(column=1, row=2, padx=10, pady=10)

        # Calculate Button
        button = ttk.Button(self, text='Calculate',
                            command=self.calc,  default='active')
        button.grid(column=0, row=4, pady=20, ipady=5, ipadx=5)

        # Copy Button
        button = ttk.Button(self, text='Copy Lay Bet',
                            command=self.copy_lay, default='active')
        button.grid(column=1, row=4,  pady=20, ipady=5, ipadx=5)

        # Calculations Frames
        bottom_frame = Frame(self)
        # Frame to display calculations
        label_frame = Frame(bottom_frame)
        tk.Label(label_frame, text='Optimal Lay Bet: ').grid(column=0, row=0, sticky='E')
        tk.Label(label_frame, text='Liability: ').grid(column=0, row=1, sticky='E')
        tk.Label(label_frame, text='Bookmaker Win Profit/Loss: ').grid(column=0, row=2, sticky='E')
        tk.Label(label_frame, text='Exchange Win Profit/Loss: ').grid(column=0, row=3, sticky='E')

        # Display Calcs Frame
        calc_frame = Frame(bottom_frame)
        self.optimal_label = tk.Label(calc_frame, text='')
        self.optimal_label.grid(column=0, row=0)
        self.liability_label = tk.Label(calc_frame, text='')
        self.liability_label.grid(column=0, row=1)
        self.back_label = tk.Label(calc_frame, text='')
        self.back_label.grid(column=0, row=2)
        self.lay_label = tk.Label(calc_frame, text='')
        self.lay_label.grid(column=0, row=3)

        label_frame.grid(column=0, row=0, sticky='W')
        calc_frame.grid(column=1, row=0, sticky='E')
        bottom_frame.grid(column=0, row=6, columnspan=2, sticky='W')

    def copy_lay(self):
        copy(self.optimal)

    def calc(self):
        bet_type = self.bet_type.get()
        stake = float(self.stake.get())
        back_odds = float(self.back_odds.get())
        lay_odds = float(self.lay_odds.get())
        comm = float(self.commission.get()) / 100

        if bet_type == 'Qualification':
            optimal = back_odds / (lay_odds - comm) * stake
            back_profit = ((back_odds - 1) * stake) - ((lay_odds - 1) * optimal)
            lay_profit = (optimal * (1 - comm)) - stake
            liability = optimal * (lay_odds - 1)

        elif bet_type == 'Free Bet (SNR)':
            optimal = (back_odds - 1) / (lay_odds - comm) * stake
            back_profit = ((back_odds - 1) * stake) - ((lay_odds - 1) * optimal)
            lay_profit = optimal * (1 - comm)
            liability = optimal * (lay_odds - 1)
        else:
            optimal = back_odds / (lay_odds - comm) * stake
            back_profit = (back_odds * stake) - ((lay_odds - 1) * optimal)
            lay_profit = optimal * (1 - comm)
            liability = optimal * (lay_odds - 1)

        self.optimal = optimal

        self.optimal_label['text'] = f'??{optimal:.2f}'
        self.liability_label['text'] = f'??{liability:.2f}'
        self.back_label['text'] = f'??{back_profit:.2f}'
        self.lay_label['text'] = f'??{lay_profit:.2f}'


if __name__ == "__main__":
    app = App()
    app.mainloop()

# TODO: Settings Page with Google Tasks Toggle
# TODO: Excel To Google Sheets
# TODO: Free Bets Dates
