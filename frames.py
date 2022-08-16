import tkinter as tk
from tkinter import ttk
from tkinter import Frame
from tkinter import messagebox
from tkcalendar import DateEntry
from datetime import date
from openpyxl import load_workbook


class Bookie(Frame):
    def __init__(self, ws):
        Frame.__init__(self, ws)
        self.ws = ws

        # Bookmaker Label and Entry
        self.label = ttk.Label(self, text='Bookmaker:')
        self.label.pack()

        self.bookie = tk.Entry(self)
        self.bookie.pack()


class Exchange(Frame):
    def __init__(self, ws):
        Frame.__init__(self, ws)
        self.ws = ws

        # Exchange Label and Entry
        self.label = ttk.Label(self, text='Exchange:')
        self.label.pack()

        # self.exchange = tk.Entry(self)
        # self.exchange.insert('end', 'Smarkets')
        # self.exchange.pack()

        self.exchange = ttk.Combobox(self, width=17)

        # Adding combobox drop down list
        self.exchange['values'] = ('Smarkets',
                                   'Matchbook',
                                   'Betfair',
                                   'Betdaq')
        self.exchange.current(0)
        self.exchange.pack()


class Details(Frame):
    def __init__(self, ws):
        Frame.__init__(self, ws)
        self.ws = ws

        # Profit Label and Entry
        self.label = ttk.Label(self, text='Details:')
        self.label.pack()

        self.details = tk.Entry(self)
        self.details.pack()


class Profit(Frame):
    def __init__(self, ws):
        Frame.__init__(self, ws)
        self.ws = ws

        # Profit Label and Entry
        self.label = ttk.Label(self, text='Profit/Loss:')
        self.label.pack()

        self.profit = tk.Entry(self)
        self.profit.pack()


class Avail(Frame):
    def __init__(self, ws):
        Frame.__init__(self, ws)
        self.ws = ws

        # Profit Label and Entry
        self.label = ttk.Label(self, text='Available From:')
        self.label.pack()

        self.cal = DateEntry(self, selectmode='day', width=17, locale='en_UK', date_pattern='dd/mm/yy')
        self.cal.pack()


class Submit(Frame):
    def __init__(self, ws):
        Frame.__init__(self, ws)
        self.ws = ws

        # Submit Button
        self.button = ttk.Button(self, text='Submit', default='active', padding=5)
        self.button.pack()


class Inputs(Frame):

    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        # Place Bookie Frame
        self.bookie = Bookie(self)
        self.bookie.grid(column=0, row=0, padx=10, pady=10)

        # Place Exchange Frame
        self.exchange = Exchange(self)
        self.exchange.grid(column=1, row=0, padx=10, pady=10)

        # Place Details Frame
        self.details = Details(self)
        self.details.grid(column=0, row=1, padx=10, pady=10)

        # Place Profit Frame
        self.profit = Profit(self)
        self.profit.grid(column=1, row=1, padx=10, pady=10)

        # Place Submit Button
        self.button = Submit(self)
        self.button.button['command'] = self.button_submit
        self.button.grid(column=0, row=3, columnspan=2, pady=10)

    def button_submit(self):

        # Read Content From File
        try:
            wb = load_workbook(filename="Betting Tool Log.xlsx")
            ws = wb['Normal']

            row = ws.max_row + 1
            get_date = date.today()
            today = get_date.strftime('%d/%m/%y')

            # Format Date Cell
            cell = ws.cell(column=2, row=row)
            cell.number_format = 'dd/mm/yy;@'
            cell.value = today

            # Format Bookie Cell
            cell = ws.cell(column=3, row=row)
            cell.value = self.bookie.bookie.get()

            # Format Exchange Cell
            cell = ws.cell(column=4, row=row)
            cell.value = self.exchange.exchange.get()

            # Format Details Cell
            cell = ws.cell(column=5, row=row)
            cell.value = self.details.details.get()

            # Format Profit Cell
            cell = ws.cell(column=6, row=row)
            cell.number_format = '£#,##0.00'
            cell.value = float(self.profit.profit.get())

            wb.save("Betting Tool Log.xlsx")
            wb.close()

        except PermissionError:
            messagebox.showerror('Error', 'Log File Permission Denied!')
            raise Exception('Log File Permission Denied!')

        self.bookie.bookie.delete(0, 'end')
        self.details.details.delete(0, 'end')
        self.profit.profit.delete(0, 'end')

        messagebox.showinfo(title='Information', message='Success')


class QualiInputs(Frame):

    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        # Place Bookie Frame
        self.bookie = Bookie(self)
        self.bookie.grid(column=0, row=0, padx=10, pady=10)

        # Place Exchange Frame
        self.exchange = Exchange(self)
        self.exchange.grid(column=1, row=0, padx=10, pady=10)

        # Place Details Frame
        self.details = Details(self)
        self.details.grid(column=0, row=1, padx=10, pady=10)

        # Place Profit Frame
        self.profit = Profit(self)
        self.profit.grid(column=1, row=1, padx=10, pady=10)

        # Place Avail Frame
        self.avail = Avail(self)
        self.avail.grid(column=0, row=2, padx=10, pady=10)

        # Place Submit Button
        self.button = Submit(self)
        self.button.button['command'] = self.button_submit
        self.button.grid(column=1, row=2, columnspan=2, pady=10)

    def button_submit(self):

        # Read Content From File
        try:
            wb = load_workbook(filename="Betting Tool Log.xlsx")
            ws = wb['Quali']

            row = ws.max_row + 1
            get_date = date.today()
            today = get_date.strftime('%d/%m/%y')

            # Format Date Cell
            cell = ws.cell(column=2, row=row)
            cell.number_format = 'dd/mm/yy;@'
            cell.value = today

            # Format Bookie Cell
            cell = ws.cell(column=3, row=row)
            cell.value = self.bookie.bookie.get()

            # Format Exchange Cell
            cell = ws.cell(column=4, row=row)
            cell.value = self.exchange.exchange.get()

            # Format Details Cell
            cell = ws.cell(column=5, row=row)
            cell.value = self.details.details.get()

            # Format Profit Cell
            cell = ws.cell(column=6, row=row)
            cell.number_format = '£#,##0.00'
            cell.value = float(self.profit.profit.get())

            # Format Profit Cell
            cell = ws.cell(column=7, row=row)
            print(self.avail.cal.get())
            cell.value = self.avail.cal.get()

            wb.save("Betting Tool Log.xlsx")
            wb.close()

        except PermissionError:
            messagebox.showerror('Error', 'Log File Permission Denied!')
            raise Exception('Log File Permission Denied!')

        self.bookie.bookie.delete(0, 'end')
        self.details.details.delete(0, 'end')
        self.profit.profit.delete(0, 'end')

        messagebox.showinfo(title='Information', message='Success')
